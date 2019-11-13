#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 11 13:11:35 2019

@author: brianmarx
"""

from openpyxl import Workbook

class excelClass(object):
    
    def __init__(self):
        pass
    
    wb = Workbook()
    
    def makeWorkbook(self, tags, title, saveTitle, ids, records):
        self.wb.title = title
        ws = self.wb.active
        ws.title = 'Data'
        
        # Label each column with the proper tag
        col = 1
        for tag in tags:
            ws.cell(column=col, row=1, value=tag)
            col += 1
        column = 1
        row = 2
        for rec in records:
            for tag in tags:
                if tag in rec and rec[tag] != '':
                    ws.cell(row=row, column=column, value=rec[tag])
                column += 1
            row += 1
            column = 1
        row += 2
        ws.cell(row=row, column=1, value='IDs:')
        row += 1
        for k, v in ids.items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
            
        self.wb.save('static/' + title + '.xlsx')